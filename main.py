import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPlainTextEdit,
    QProgressDialog,
    QPushButton,
    QSplitter,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)


class CreateFolderApp(QWidget):
    PATH_PLACEHOLDER = "尚未選擇路徑"
    PREVIEW_LIMIT = 5
    CONNECTOR_DEPTH = 6

    def __init__(self):
        super().__init__()
        self.groups = []
        self.current_group_index = None
        self._loading_group = False
        self.preview_compact_mode = True
        self.init_ui()
        self.add_group()

    def init_ui(self):
        self.setWindowTitle("資料夾建立工具 v4.5")
        self.resize(1450, 900)

        self.setStyleSheet(
            """
            QWidget { background-color: #1e1e1e; color: #dcdcdc; font-size: 14px; }
            QLabel { color: #569cd6; font-weight: bold; }
            QLineEdit, QPlainTextEdit, QListWidget { background-color: #3c3c3c; border: 1px solid #555; color: white; padding: 5px; }
            QPushButton { background-color: #333; border: 1px solid #555; padding: 8px; min-width: 80px; }
            QPushButton:hover { background-color: #444; border-color: #0078d7; }
            QGroupBox { border: 1px solid #555; margin-top: 10px; padding-top: 8px; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; color: #9cdcfe; }
            """
        )

        main_layout = QVBoxLayout(self)

        top_nav = QHBoxLayout()
        path_group = QVBoxLayout()
        self.btn_select = QPushButton("選擇輸出路徑")
        self.lbl_path = QLabel(self.PATH_PLACEHOLDER)
        self.lbl_path.setStyleSheet("background: #073642; color: #268bd2; padding: 8px; border-radius: 4px;")
        path_group.addWidget(self.btn_select)
        path_group.addWidget(self.lbl_path)

        param_grid = QGridLayout()
        param_grid.addWidget(QLabel("DUT Name:"), 0, 0)
        self.edit_dut = QLineEdit()
        param_grid.addWidget(self.edit_dut, 0, 1)
        param_grid.addWidget(QLabel("Version:"), 1, 0)
        self.edit_ver = QLineEdit()
        param_grid.addWidget(self.edit_ver, 1, 1)
        param_grid.addWidget(QLabel("主板板廠清單"), 2, 0)
        self.txt_mainboard_vendors = QPlainTextEdit()
        self.txt_mainboard_vendors.setPlaceholderText("每行一個主板板廠")
        self.txt_mainboard_vendors.setFixedHeight(90)
        param_grid.addWidget(self.txt_mainboard_vendors, 2, 1)

        top_nav.addLayout(path_group, 3)
        top_nav.addSpacing(20)
        top_nav.addLayout(param_grid, 2)
        main_layout.addLayout(top_nav)

        toolbar = QHBoxLayout()
        self.btn_import_json = QPushButton("匯入 JSON")
        self.btn_export_json = QPushButton("匯出 JSON")
        self.btn_export_xlsx = QPushButton("匯出測試清單 XLSX")
        toolbar.addWidget(self.btn_import_json)
        toolbar.addWidget(self.btn_export_json)
        toolbar.addWidget(self.btn_export_xlsx)
        toolbar.addStretch()
        main_layout.addLayout(toolbar)

        center_splitter = QSplitter(Qt.Horizontal)

        left_box = QGroupBox("設定列表")
        left_layout = QVBoxLayout(left_box)
        self.list_groups = QListWidget()
        left_layout.addWidget(self.list_groups)

        left_btns_row1 = QHBoxLayout()
        left_btns_row2 = QHBoxLayout()
        self.btn_add_group = QPushButton("新增設定")
        self.btn_copy_group = QPushButton("複製設定")
        self.btn_del_group = QPushButton("刪除設定")
        self.btn_clear_current_group = QPushButton("清空目前設定")
        self.btn_clear_all_groups = QPushButton("清空全部設定")
        left_btns_row1.addWidget(self.btn_add_group)
        left_btns_row1.addWidget(self.btn_copy_group)
        left_btns_row1.addWidget(self.btn_del_group)
        left_btns_row2.addWidget(self.btn_clear_current_group)
        left_btns_row2.addWidget(self.btn_clear_all_groups)
        left_layout.addLayout(left_btns_row1)
        left_layout.addLayout(left_btns_row2)

        setting_box = QGroupBox("設定內容")
        form = QFormLayout(setting_box)

        self.edit_card = QLineEdit()
        self.txt_vendors = QPlainTextEdit()
        self.txt_cables = QPlainTextEdit()
        self.txt_connectors = QPlainTextEdit()
        self.edit_slot_prefix = QLineEdit()
        self.edit_slot_list = QLineEdit()
        self.edit_gen_list = QLineEdit()
        self.edit_slot_list.setPlaceholderText("可空白，例如 1-5 或 2,4")
        self.edit_gen_list.setPlaceholderText("可空白，例如 2,4 或 1,3-5,8")

        self.txt_vendors.setPlaceholderText("每行一個小卡板廠")
        self.txt_cables.setPlaceholderText("每行一個 Cable")
        self.txt_connectors.setPlaceholderText("每行一個 Connector")
        self.txt_vendors.setFixedHeight(80)
        self.txt_cables.setFixedHeight(80)
        self.txt_connectors.setFixedHeight(80)

        form.addRow("小卡名稱（可空白）", self.edit_card)
        form.addRow("小卡板廠清單", self.txt_vendors)
        form.addRow("Cable廠商清單", self.txt_cables)
        form.addRow("主板Connector名稱", self.txt_connectors)
        form.addRow("(小卡)Slot 名稱", self.edit_slot_prefix)
        form.addRow("Slot 編號", self.edit_slot_list)
        form.addRow("GEN 編號", self.edit_gen_list)
        self.lbl_number_rule_hint = QLabel(
            "編號規則說明（適用於 Slot / GEN）：\n"
            "1-5 代表連續範圍：1,2,3,4,5\n"
            "1,5 代表只建立 1 和 5\n"
            "1,3-5,8 代表：1,3,4,5,8\n"
            "GEN 編號可空白，空白時不建立 GEN 資料夾"
        )
        self.lbl_number_rule_hint.setWordWrap(True)
        self.lbl_number_rule_hint.setStyleSheet("color: #888; font-size: 12px; padding-left: 4px;")
        form.addRow("", self.lbl_number_rule_hint)

        preview_box = QGroupBox("資料夾結構預覽")
        preview_layout = QVBoxLayout(preview_box)
        self.lbl_preview_stats = QLabel("預估建立：Slot 資料夾 0 個，GEN 資料夾 0 個")
        self.lbl_preview_mode = QLabel("目前模式：精簡模式")
        self.lbl_preview_warning = QLabel("")
        self.lbl_preview_warning.setStyleSheet("color: #d19a66;")
        preview_layout.addWidget(self.lbl_preview_stats)
        preview_layout.addWidget(self.lbl_preview_mode)
        preview_layout.addWidget(self.lbl_preview_warning)
        preview_btns = QHBoxLayout()
        self.btn_expand_all = QPushButton("詳細模式")
        self.btn_collapse_all = QPushButton("精簡模式")
        self.btn_collapse_all.setEnabled(False)
        preview_btns.addWidget(self.btn_expand_all)
        preview_btns.addWidget(self.btn_collapse_all)
        preview_btns.addStretch()
        preview_layout.addLayout(preview_btns)
        self.preview_tree = QTreeWidget()
        self.preview_tree.setHeaderHidden(True)
        preview_layout.addWidget(self.preview_tree)

        center_splitter.addWidget(left_box)
        center_splitter.addWidget(setting_box)
        center_splitter.addWidget(preview_box)
        center_splitter.setSizes([280, 520, 650])
        center_splitter.setStretchFactor(0, 0)
        center_splitter.setStretchFactor(1, 1)
        center_splitter.setStretchFactor(2, 2)
        main_layout.addWidget(center_splitter, 1)

        bottom_row = QHBoxLayout()
        self.btn_confirm = QPushButton("開始建立資料夾")
        self.btn_confirm.setFixedHeight(45)
        self.btn_confirm.setStyleSheet("background-color: #0e639c; font-size: 16px; font-weight: bold;")
        bottom_row.addStretch()
        bottom_row.addWidget(self.btn_confirm, 1)
        main_layout.addLayout(bottom_row)

        self.btn_select.clicked.connect(self.select_directory)
        self.btn_import_json.clicked.connect(self.import_json)
        self.btn_export_json.clicked.connect(self.export_json)
        self.btn_export_xlsx.clicked.connect(self.export_test_list_xlsx)
        self.btn_confirm.clicked.connect(self.process_data)

        self.btn_add_group.clicked.connect(self.add_group)
        self.btn_copy_group.clicked.connect(self.copy_group)
        self.btn_del_group.clicked.connect(self.delete_group)
        self.btn_clear_current_group.clicked.connect(self.clear_current_group)
        self.btn_clear_all_groups.clicked.connect(self.clear_all_groups)
        self.list_groups.currentRowChanged.connect(self.on_group_selected)

        self.edit_dut.editingFinished.connect(self.update_preview)
        self.edit_ver.editingFinished.connect(self.update_preview)

        self.edit_card.textChanged.connect(self.save_current_group)
        self.edit_slot_prefix.textChanged.connect(self.save_current_group)
        self.edit_slot_list.textChanged.connect(self.save_current_group)
        self.edit_gen_list.textChanged.connect(self.save_current_group)
        self.txt_vendors.textChanged.connect(self.save_current_group)
        self.txt_cables.textChanged.connect(self.save_current_group)
        self.txt_connectors.textChanged.connect(self.save_current_group)
        self.txt_mainboard_vendors.textChanged.connect(self.update_preview)
        self.btn_expand_all.clicked.connect(self._set_detailed_preview)
        self.btn_collapse_all.clicked.connect(self._set_compact_preview)

    def sanitize_name(self, name):
        if not name:
            return ""
        return re.sub(r'[\\/:*?"<>|]', "_", name).strip()

    def _parse_multiline(self, text):
        values = []
        for token in re.split(r"[\n,;、]+", text):
            clean = self.sanitize_name(token.strip())
            if clean and clean not in values:
                values.append(clean)
        return values

    def _new_group(self):
        return {
            "card": "",
            "vendors": [],
            "cables": [],
            "connectors": [],
            "slot_prefix": "",
            "slot_list": "",
            "gen_list": "",
        }

    def _refresh_group_list(self):
        self.list_groups.blockSignals(True)
        self.list_groups.clear()
        for i, g in enumerate(self.groups):
            label = self._group_display_label(g, i)
            self.list_groups.addItem(QListWidgetItem(label))
        self.list_groups.blockSignals(False)

    def _update_group_list_label(self, idx):
        if idx is None or idx < 0 or idx >= len(self.groups):
            return
        item = self.list_groups.item(idx)
        if not item:
            return
        g = self.groups[idx]
        label = self._group_display_label(g, idx)
        item.setText(label)

    def _group_display_label(self, group, index):
        card = self.sanitize_name(group.get("card", "")).strip()
        slot_prefix = self.sanitize_name(group.get("slot_prefix", "")).strip()
        if card:
            return card
        if slot_prefix:
            return f"Slot: {slot_prefix}"
        return f"未命名設定 {index + 1}"

    def add_group(self):
        self.save_current_group()
        self.groups.append(self._new_group())
        self._refresh_group_list()
        self.list_groups.setCurrentRow(len(self.groups) - 1)

    def copy_group(self):
        if self.current_group_index is None or self.current_group_index < 0:
            return
        self.save_current_group()
        src = self.groups[self.current_group_index]
        copied = {
            "card": src.get("card", ""),
            "vendors": list(src.get("vendors", [])),
            "cables": list(src.get("cables", [])),
            "connectors": list(src.get("connectors", [])),
            "slot_prefix": src.get("slot_prefix", ""),
            "slot_list": src.get("slot_list", ""),
            "gen_list": src.get("gen_list", ""),
        }
        self.groups.append(copied)
        self._refresh_group_list()
        self.list_groups.setCurrentRow(len(self.groups) - 1)

    def delete_group(self):
        if self.current_group_index is None or self.current_group_index < 0:
            return
        reply = QMessageBox.question(self, "確認", "確定要刪除目前設定嗎？")
        if reply != QMessageBox.StandardButton.Yes:
            return
        del self.groups[self.current_group_index]
        if not self.groups:
            self.groups.append(self._new_group())
        self._refresh_group_list()
        self.list_groups.setCurrentRow(min(self.current_group_index, len(self.groups) - 1))

    def clear_current_group(self):
        if self.current_group_index is None or self.current_group_index < 0 or self.current_group_index >= len(self.groups):
            return
        reply = QMessageBox.question(self, "確認", "確定要清空目前設定嗎？")
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.groups[self.current_group_index] = self._new_group()
        self.load_current_group()
        self._update_group_list_label(self.current_group_index)
        self.update_preview()

    def clear_all_groups(self):
        reply = QMessageBox.question(self, "確認", "確定要清空全部設定嗎？")
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.groups = [self._new_group()]
        self._refresh_group_list()
        self.list_groups.setCurrentRow(0)
        self.update_preview()

    def on_group_selected(self, row):
        self.current_group_index = row
        self.load_current_group()

    def load_current_group(self):
        if self.current_group_index is None or self.current_group_index < 0 or self.current_group_index >= len(self.groups):
            return

        g = self.groups[self.current_group_index]
        self._loading_group = True
        self.edit_card.setText(g.get("card", ""))
        self.txt_vendors.setPlainText("\n".join(g.get("vendors", [])))
        self.txt_cables.setPlainText("\n".join(g.get("cables", [])))
        self.txt_connectors.setPlainText("\n".join(g.get("connectors", [])))
        self.edit_slot_prefix.setText(g.get("slot_prefix", ""))
        self.edit_slot_list.setText(g.get("slot_list", ""))
        self.edit_gen_list.setText(g.get("gen_list", ""))
        self._loading_group = False
        self.update_preview()

    def save_current_group(self):
        if self._loading_group:
            return
        if self.current_group_index is None or self.current_group_index < 0 or self.current_group_index >= len(self.groups):
            return

        g = self.groups[self.current_group_index]
        g["card"] = self.sanitize_name(self.edit_card.text().strip())
        g["vendors"] = self._parse_multiline(self.txt_vendors.toPlainText())
        g["cables"] = self._parse_multiline(self.txt_cables.toPlainText())
        g["connectors"] = self._parse_multiline(self.txt_connectors.toPlainText())
        g["slot_prefix"] = self.sanitize_name(self.edit_slot_prefix.text().strip())
        g["slot_list"] = self.edit_slot_list.text().strip()
        g["gen_list"] = self.edit_gen_list.text().strip()

        self._update_group_list_label(self.current_group_index)
        self.update_preview()

    def _validated_groups(self, include_errors=False):
        valid = []
        errors = []
        for idx, g in enumerate(self.groups):
            slot_prefix = self.sanitize_name(str(g.get("slot_prefix", "")).strip())
            slot_list_text = str(g.get("slot_list", "")).strip()
            gen_list_text = str(g.get("gen_list", "")).strip()
            display_name = self._group_display_label(g, idx)

            if not slot_prefix:
                if any([slot_list_text, gen_list_text]):
                    errors.append(f"{display_name}: Slot 名稱必填")
                continue

            try:
                slot_numbers = self._parse_slot_numbers(slot_list_text)
            except ValueError as ve:
                msg = str(ve)
                if "反向" in msg or "Slot 編號必填" in msg:
                    errors.append(f"{display_name}: {msg}")
                else:
                    errors.append(f"{display_name}: Slot 編號格式錯誤")
                continue
            try:
                gen_numbers = self._parse_gen_numbers(gen_list_text)
            except ValueError as ve:
                msg = str(ve)
                if "反向" in msg:
                    errors.append(f"{display_name}: {msg}")
                else:
                    errors.append(f"{display_name}: GEN 編號格式錯誤")
                continue

            valid.append(
                {
                    "index": idx,
                    "card": self.sanitize_name(g.get("card", "")),
                    "vendors": list(g.get("vendors", [])),
                    "cables": list(g.get("cables", [])),
                    "connectors": list(g.get("connectors", [])),
                    "slot_prefix": slot_prefix,
                    "slot_list": slot_list_text,
                    "gen_list": gen_list_text,
                    "slot_numbers": slot_numbers,
                    "slot_names": (
                        [f"{slot_prefix}{'_' if slot_prefix and slot_prefix[-1].isdigit() else ''}{n}" for n in slot_numbers]
                        if slot_numbers else [slot_prefix]
                    ),
                    "gen_numbers": gen_numbers,
                    "gen_start": min(gen_numbers) if gen_numbers else None,
                    "gen_end": max(gen_numbers) if gen_numbers else None,
                }
            )
        if include_errors:
            return valid, errors
        return valid

    def _parse_number_list(self, text, label):
        if not text.strip():
            return []
        if re.search(r"[^0-9,\-;、\s]", text):
            raise ValueError(f"{label}只允許數字、逗號、分號、頓號、換行、dash")

        numbers = []
        seen = set()
        parts = re.split(r"[,;、\n]+", text)
        for raw in parts:
            token = raw.strip()
            if not token:
                continue
            if "-" in token:
                m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", token)
                if not m:
                    raise ValueError(f"{label}範圍格式錯誤：{token}")
                s_num = int(m.group(1))
                e_num = int(m.group(2))
                if e_num < s_num:
                    raise ValueError(f"{label}範圍不可反向：{token}")
                for n in range(s_num, e_num + 1):
                    if n not in seen:
                        seen.add(n)
                        numbers.append(n)
            else:
                if not token.isdigit():
                    raise ValueError(f"{label}格式錯誤：{token}")
                n = int(token)
                if n not in seen:
                    seen.add(n)
                    numbers.append(n)
        if not numbers:
            raise ValueError(f"{label}解析後不可為空")
        return sorted(numbers)

    def _parse_slot_numbers(self, slot_list_text):
        if slot_list_text.strip():
            return self._parse_number_list(slot_list_text, "Slot 編號")
        return []

    def _parse_gen_numbers(self, gen_list_text):
        if gen_list_text.strip():
            return self._parse_number_list(gen_list_text, "GEN 編號")
        return []

    def _preview_groups(self):
        preview_groups = []
        errors = []
        for idx, g in enumerate(self.groups):
            display_name = self._group_display_label(g, idx)
            slot_prefix = self.sanitize_name(str(g.get("slot_prefix", "")).strip())
            if not slot_prefix:
                continue

            slot_list_text = str(g.get("slot_list", "")).strip()
            gen_list_text = str(g.get("gen_list", "")).strip()
            try:
                slot_numbers = self._parse_slot_numbers(slot_list_text)
            except ValueError as ve:
                msg = str(ve)
                if "反向" in msg or "Slot 編號必填" in msg:
                    errors.append(f"{display_name}: {msg}")
                else:
                    errors.append(f"{display_name}: Slot 編號格式錯誤")
                continue

            try:
                gen_numbers = self._parse_gen_numbers(gen_list_text)
            except ValueError as ve:
                msg = str(ve)
                if "反向" in msg or "不可只填一個" in msg:
                    errors.append(f"{display_name}: {msg}")
                else:
                    errors.append(f"{display_name}: GEN 編號格式錯誤")
                continue

            preview_groups.append(
                {
                    "index": idx,
                    "card": self.sanitize_name(g.get("card", "")),
                    "vendors": list(g.get("vendors", [])),
                    "cables": list(g.get("cables", [])),
                    "connectors": list(g.get("connectors", [])),
                    "slot_prefix": slot_prefix,
                    "slot_list": slot_list_text,
                    "gen_list": gen_list_text,
                    "slot_numbers": slot_numbers,
                    "slot_names": (
                        [f"{slot_prefix}{'_' if slot_prefix and slot_prefix[-1].isdigit() else ''}{n}" for n in slot_numbers]
                        if slot_numbers else [slot_prefix]
                    ),
                    "gen_numbers": gen_numbers,
                    "gen_start": min(gen_numbers) if gen_numbers else None,
                    "gen_end": max(gen_numbers) if gen_numbers else None,
                }
            )
        return preview_groups, errors

    def _looks_like_numbered_slot(self, slot_prefix):
        return slot_prefix.upper() in ["PCIE", "M2", "SLOT"]

    def _confirm_blank_slot_numbers(self, groups, action_name):
        warnings = []
        for cfg in groups:
            if not cfg["slot_numbers"] and self._looks_like_numbered_slot(cfg["slot_prefix"]):
                warnings.append(f"- {self._group_display_label(self.groups[cfg['index']], cfg['index'])}: Slot: {cfg['slot_prefix']}")
        if not warnings:
            return True
        msg = (
            "以下設定 Slot 編號空白，將直接建立固定名稱資料夾：\n"
            + "\n".join(warnings[:10])
            + f"\n是否繼續{action_name}？"
        )
        reply = QMessageBox.question(self, "確認", msg)
        return reply == QMessageBox.StandardButton.Yes

    def _ensure_child(self, parent_item, text):
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            if child.text(0) == text:
                return child
        child = QTreeWidgetItem([text])
        parent_item.addChild(child)
        return child

    def _add_limited_paths(self, root_item, paths, compact_mode=True):
        children_map = {}
        for path in paths:
            if not path:
                continue
            children_map.setdefault(path[0], []).append(path[1:])

        names = sorted(children_map.keys(), key=self._natural_key)
        if compact_mode:
            compact_label = self._compact_range_label(names)
            if compact_label:
                # 只有在所有被壓縮節點的下一層結構一致時，才允許壓縮父層
                # 例如：PCIE1~PCIE5 底下都同樣是 GEN1~GEN5
                child_signatures = [self._path_signature(children_map[name]) for name in names]
                if len(set(child_signatures)) == 1:
                    compact_node = self._ensure_child(root_item, compact_label)
                    merged_children = []
                    for name in names:
                        merged_children.extend([p for p in children_map[name] if p])
                    self._add_limited_paths(compact_node, merged_children, compact_mode=compact_mode)
                    return

        display_names = names[: self.PREVIEW_LIMIT] if compact_mode else names
        for name in display_names:
            child = self._ensure_child(root_item, name)
            remain_paths = [p for p in children_map[name] if p]
            self._add_limited_paths(child, remain_paths, compact_mode=compact_mode)
        if compact_mode and len(names) > self.PREVIEW_LIMIT:
            self._ensure_child(root_item, f"... 另有 {len(names) - self.PREVIEW_LIMIT} 個")

    def _natural_key(self, text):
        m = re.match(r"^(.*?)(\d+)$", text or "")
        if not m:
            return (text, -1)
        return (m.group(1), int(m.group(2)))

    def _compact_range_label(self, names):
        if len(names) < 2:
            return None
        parsed = []
        for name in names:
            m = re.match(r"^([A-Za-z_]+)(\d+)$", name)
            if not m:
                return None
            parsed.append((m.group(1), int(m.group(2))))
        prefixes = {p for p, _ in parsed}
        if len(prefixes) != 1:
            return None
        nums = sorted(n for _, n in parsed)
        if nums != list(range(nums[0], nums[-1] + 1)):
            return None
        prefix = parsed[0][0]
        if not (prefix.upper().startswith("GEN") or prefix.upper().startswith("PCIE") or prefix.upper().startswith("SLOT")):
            return None
        return f"{prefix}{nums[0]} ~ {prefix}{nums[-1]}"

    def _path_signature(self, paths):
        normalized = []
        for p in paths:
            normalized.append(tuple(p))
        normalized.sort()
        return tuple(normalized)

    def _estimate_folder_counts(self):
        valid_groups, _ = self._preview_groups()
        mainboard_vendors = self._parse_multiline(self.txt_mainboard_vendors.toPlainText())
        mainboard_vendors = mainboard_vendors if mainboard_vendors else [""]

        slot_total = 0
        gen_total = 0
        for cfg in valid_groups:
            mb_cnt = len(mainboard_vendors)
            v_cnt = len(cfg["vendors"]) if cfg["vendors"] else 1
            c_cnt = len(cfg["cables"]) if cfg["cables"] else 1
            k_cnt = len(cfg["connectors"]) if cfg["connectors"] else 1
            slot_cnt = len(cfg["slot_names"])
            combo_cnt = mb_cnt * v_cnt * c_cnt * k_cnt
            slot_total += combo_cnt * slot_cnt
            gen_total += combo_cnt * slot_cnt * len(cfg["gen_numbers"])
        return slot_total, gen_total

    def _set_expand_level(self, item, depth):
        item.setExpanded(depth <= self.CONNECTOR_DEPTH)
        for i in range(item.childCount()):
            self._set_expand_level(item.child(i), depth + 1)

    def _collapse_preview_to_connector(self):
        for i in range(self.preview_tree.topLevelItemCount()):
            self._set_expand_level(self.preview_tree.topLevelItem(i), 0)

    def _set_detailed_preview(self):
        self.preview_compact_mode = False
        self.lbl_preview_mode.setText("目前模式：詳細模式")
        self.btn_expand_all.setEnabled(False)
        self.btn_collapse_all.setEnabled(True)
        self.update_preview()
        self.preview_tree.expandAll()

    def _set_compact_preview(self):
        self.preview_compact_mode = True
        self.lbl_preview_mode.setText("目前模式：精簡模式")
        self.btn_collapse_all.setEnabled(False)
        self.btn_expand_all.setEnabled(True)
        self.update_preview()

    def _group_expanded_paths(self, cfg):
        paths = []
        mainboard_vendors = self._parse_multiline(self.txt_mainboard_vendors.toPlainText())
        mainboard_vendors = mainboard_vendors if mainboard_vendors else [""]
        vendors = cfg["vendors"] if cfg["vendors"] else [""]
        cables = cfg["cables"] if cfg["cables"] else [""]
        connectors = cfg["connectors"] if cfg["connectors"] else [""]
        joiner = "_" if cfg["slot_prefix"] and cfg["slot_prefix"][-1].isdigit() else ""

        for mb in mainboard_vendors:
            for v in vendors:
                for c in cables:
                    for k in connectors:
                        base = [p for p in [mb, cfg["card"], v, c, k] if p]
                        for slot_name in cfg["slot_names"]:
                            slot_path = base + [slot_name]
                            if not cfg["gen_numbers"]:
                                paths.append(slot_path)
                            else:
                                for g in cfg["gen_numbers"]:
                                    paths.append(slot_path + [f"GEN{g}"])
        return paths

    def update_preview(self):
        self.preview_tree.clear()

        dut = self.sanitize_name(self.edit_dut.text().strip()) or "DUT"
        ver = self.sanitize_name(self.edit_ver.text().strip())

        root_item = QTreeWidgetItem([dut])
        self.preview_tree.addTopLevelItem(root_item)

        base_item = root_item
        if ver:
            base_item = self._ensure_child(root_item, ver)

        preview_groups, preview_errors = self._preview_groups()
        all_paths = []
        for cfg in preview_groups:
            all_paths.extend(self._group_expanded_paths(cfg))

        self._add_limited_paths(base_item, all_paths, compact_mode=self.preview_compact_mode)
        slot_total, gen_total = self._estimate_folder_counts()
        self.lbl_preview_stats.setText(f"預估建立：Slot 資料夾 {slot_total} 個，GEN 資料夾 {gen_total} 個")
        if preview_errors:
            lines = "\n".join(f"- {e}" for e in preview_errors[:3])
            self.lbl_preview_warning.setText("Preview 警告：\n" + lines)
        else:
            self.lbl_preview_warning.setText("")
        if self.preview_compact_mode:
            self._collapse_preview_to_connector()
        else:
            self.preview_tree.expandAll()

    def select_directory(self):
        folder_path = QFileDialog.getExistingDirectory(self, "選擇輸出資料夾")
        if folder_path:
            self.lbl_path.setText(folder_path)

    def export_json(self):
        self.save_current_group()
        path, _ = QFileDialog.getSaveFileName(self, "匯出 JSON", "", "JSON Files (*.json)")
        if not path:
            return

        payload_groups = []
        for g in self.groups:
            payload_groups.append(
                {
                    "card": self.sanitize_name(g.get("card", "")),
                    "vendors": list(g.get("vendors", [])),
                    "cables": list(g.get("cables", [])),
                    "connectors": list(g.get("connectors", [])),
                    "slot_prefix": self.sanitize_name(g.get("slot_prefix", "")),
                    "slot_list": str(g.get("slot_list", "")).strip(),
                    "gen_list": str(g.get("gen_list", "")).strip(),
                }
            )

        data = {
            "dut": self.sanitize_name(self.edit_dut.text().strip()),
            "version": self.sanitize_name(self.edit_ver.text().strip()),
            "output_path": "" if self.lbl_path.text() == self.PATH_PLACEHOLDER else self.lbl_path.text(),
            "mainboard_vendors": self._parse_multiline(self.txt_mainboard_vendors.toPlainText()),
            "groups": payload_groups,
        }

        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "完成", "JSON 匯出完成")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def import_json(self):
        path, _ = QFileDialog.getOpenFileName(self, "匯入 JSON", "", "JSON Files (*.json)")
        if not path:
            return

        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.edit_dut.setText(self.sanitize_name(str(data.get("dut", ""))))
            self.edit_ver.setText(self.sanitize_name(str(data.get("version", ""))))
            output_path = str(data.get("output_path", "")).strip()
            self.lbl_path.setText(output_path if output_path else self.PATH_PLACEHOLDER)
            self.txt_mainboard_vendors.setPlainText("\n".join(
                [self.sanitize_name(str(v)) for v in data.get("mainboard_vendors", []) if self.sanitize_name(str(v))]
            ))

            self.groups = []
            for g in data.get("groups", []):
                slot_list = str(g.get("slot_list", "")).strip()
                gen_list = str(g.get("gen_list", "")).strip()
                if not slot_list:
                    s_start = g.get("slot_start", "")
                    s_end = g.get("slot_end", "")
                    if str(s_start).strip() != "" and str(s_end).strip() != "":
                        slot_list = f"{s_start}-{s_end}"
                if not gen_list:
                    g_start = g.get("gen_start", "")
                    g_end = g.get("gen_end", "")
                    if str(g_start).strip() != "" and str(g_end).strip() != "":
                        gen_list = f"{g_start}-{g_end}"
                self.groups.append(
                    {
                        "card": self.sanitize_name(str(g.get("card", ""))),
                        "vendors": [self.sanitize_name(str(v)) for v in g.get("vendors", []) if self.sanitize_name(str(v))],
                        "cables": [self.sanitize_name(str(v)) for v in g.get("cables", []) if self.sanitize_name(str(v))],
                        "connectors": [self.sanitize_name(str(v)) for v in g.get("connectors", []) if self.sanitize_name(str(v))],
                        "slot_prefix": self.sanitize_name(str(g.get("slot_prefix", ""))),
                        "slot_list": slot_list,
                        "gen_list": gen_list,
                    }
                )

            if not self.groups:
                self.groups.append(self._new_group())

            self._refresh_group_list()
            self.list_groups.setCurrentRow(0)
            self.update_preview()
            QMessageBox.information(self, "完成", "JSON 匯入完成")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def _test_list_rows(self):
        rows = []
        valid_groups = self._validated_groups()
        mainboard_vendors = self._parse_multiline(self.txt_mainboard_vendors.toPlainText())
        mainboard_vendors = mainboard_vendors if mainboard_vendors else [""]

        for cfg in valid_groups:
            vendors = cfg["vendors"] if cfg["vendors"] else [""]
            cables = cfg["cables"] if cfg["cables"] else [""]
            connectors = cfg["connectors"] if cfg["connectors"] else [""]
            joiner = "_" if cfg["slot_prefix"] and cfg["slot_prefix"][-1].isdigit() else ""

            for mb in mainboard_vendors:
                for v in vendors:
                    for c in cables:
                        for k in connectors:
                            for slot_name in cfg["slot_names"]:
                                if not cfg["gen_numbers"]:
                                    rows.append([mb, cfg["card"], v, c, k, slot_name, "", "", "", "", ""])
                                else:
                                    for g in cfg["gen_numbers"]:
                                        gen_name = f"GEN{g}"
                                        rows.append([mb, cfg["card"], v, c, k, slot_name, gen_name, "", "", "", ""])
        return rows

    def export_test_list_xlsx(self):
        self.save_current_group()
        path, _ = QFileDialog.getSaveFileName(self, "匯出測試清單 XLSX", "", "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        _, errors = self._validated_groups(include_errors=True)
        if errors:
            return QMessageBox.warning(self, "輸入有誤", "請先修正以下問題：\n" + "\n".join(errors[:10]))
        valid_groups = self._validated_groups()
        if not self._confirm_blank_slot_numbers(valid_groups, "匯出"):
            return

        rows = self._test_list_rows()
        if not rows:
            return QMessageBox.information(self, "提醒", "目前沒有可匯出的有效測試資料")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Matrix"

            headers = [
                "Mainboard Vendor",
                "Card",
                "Card Vendor",
                "Cable",
                "Connector",
                "Slot",
                "GEN",
                "Result",
                "Tester",
                "Date",
                "Note",
            ]
            ws.append(headers)
            for row in rows:
                ws.append(row)

            for cell in ws[1]:
                cell.font = Font(bold=True)

            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            ws.freeze_panes = "A2"

            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(headers[col_idx - 1])
                for cell in ws[col_letter]:
                    if cell.value is None:
                        continue
                    max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

            result_col = "H"
            dv = DataValidation(type="list", formula1='"PASS,FAIL,NA,BLOCK"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{result_col}2:{result_col}{ws.max_row}")

            pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.conditional_formatting.add(
                f"{result_col}2:{result_col}{ws.max_row}",
                FormulaRule(formula=[f'${result_col}2="PASS"'], fill=pass_fill),
            )
            ws.conditional_formatting.add(
                f"{result_col}2:{result_col}{ws.max_row}",
                FormulaRule(formula=[f'${result_col}2="FAIL"'], fill=fail_fill),
            )

            summary = wb.create_sheet("Summary")
            summary["A1"] = "Total Cases"
            summary["B1"] = ws.max_row - 1
            summary["A2"] = "PASS Count"
            summary["B2"] = '=COUNTIF(\'Test Matrix\'!H:H,"PASS")'
            summary["A3"] = "FAIL Count"
            summary["B3"] = '=COUNTIF(\'Test Matrix\'!H:H,"FAIL")'
            summary["A4"] = "PASS Rate"
            summary["B4"] = '=IF(B1=0,0,B2/B1)'
            summary["B4"].number_format = "0.00%"
            for c in ["A1", "A2", "A3", "A4"]:
                summary[c].font = Font(bold=True)
            summary.column_dimensions["A"].width = 18
            summary.column_dimensions["B"].width = 18

            wb.save(path)
            QMessageBox.information(self, "完成", "測試清單 XLSX 匯出完成")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def process_data(self):
        self.save_current_group()

        root = self.lbl_path.text().strip()
        dut = self.sanitize_name(self.edit_dut.text().strip())
        ver = self.sanitize_name(self.edit_ver.text().strip())

        if root == self.PATH_PLACEHOLDER or not dut:
            return QMessageBox.warning(self, "提醒", "請先選擇輸出路徑，並填寫 DUT Name")

        valid_groups, errors = self._validated_groups(include_errors=True)
        if errors:
            return QMessageBox.warning(self, "輸入有誤", "請先修正以下問題：\n" + "\n".join(errors[:10]))
        if not valid_groups:
            return QMessageBox.information(self, "提醒", "請至少填寫一個有效群組")
        if not self._confirm_blank_slot_numbers(valid_groups, "建立"):
            return

        base_path = Path(root) / dut
        if ver:
            base_path = base_path / ver

        total = 0
        mainboard_vendors = self._parse_multiline(self.txt_mainboard_vendors.toPlainText())
        mainboard_vendors = mainboard_vendors if mainboard_vendors else [""]
        for cfg in valid_groups:
            mb_cnt = len(mainboard_vendors)
            v_cnt = len(cfg["vendors"]) if cfg["vendors"] else 1
            c_cnt = len(cfg["cables"]) if cfg["cables"] else 1
            k_cnt = len(cfg["connectors"]) if cfg["connectors"] else 1
            s_cnt = len(cfg["slot_names"])
            if cfg["gen_numbers"]:
                total += mb_cnt * v_cnt * c_cnt * k_cnt * s_cnt * len(cfg["gen_numbers"])
            else:
                total += mb_cnt * v_cnt * c_cnt * k_cnt * s_cnt

        progress = QProgressDialog("正在建立資料夾...", "取消", 0, max(1, total), self)
        progress.setWindowModality(Qt.WindowModal)

        try:
            step = 0
            for cfg in valid_groups:
                mainboard_vendors = self._parse_multiline(self.txt_mainboard_vendors.toPlainText())
                mainboard_vendors = mainboard_vendors if mainboard_vendors else [""]
                vendors = cfg["vendors"] if cfg["vendors"] else [""]
                cables = cfg["cables"] if cfg["cables"] else [""]
                connectors = cfg["connectors"] if cfg["connectors"] else [""]
                for mb in mainboard_vendors:
                    for v in vendors:
                        for c in cables:
                            for k in connectors:
                                parent = base_path
                                for p in [mb, cfg["card"], v, c, k]:
                                    if p:
                                        parent = parent / p

                                for slot_name in cfg["slot_names"]:
                                    if progress.wasCanceled():
                                        break

                                    slot_dir = parent / slot_name
                                    slot_dir.mkdir(parents=True, exist_ok=True)

                                    if cfg["gen_numbers"]:
                                        for g in cfg["gen_numbers"]:
                                            (slot_dir / f"GEN{g}").mkdir(parents=True, exist_ok=True)
                                            step += 1
                                            progress.setValue(step)
                                    else:
                                        step += 1
                                        progress.setValue(step)

                                if progress.wasCanceled():
                                    break
                            if progress.wasCanceled():
                                break
                        if progress.wasCanceled():
                            break
                    if progress.wasCanceled():
                        break
                if progress.wasCanceled():
                    break

            if progress.wasCanceled():
                QMessageBox.information(self, "完成", "作業已取消（已建立部分資料夾）")
            else:
                reply = QMessageBox.question(
                    self,
                    "完成",
                    "資料夾建立完成，是否開啟 DUT 資料夾？"
                )
                if reply == QMessageBox.StandardButton.Yes:
                    open_path = Path(root) / dut
                    try:
                        os.startfile(str(open_path))
                    except Exception as open_err:
                        QMessageBox.warning(self, "提醒", f"無法開啟資料夾：{open_err}")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"建立失敗: {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CreateFolderApp()
    window.show()
    sys.exit(app.exec())
